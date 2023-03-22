from openpyxl import *
import random
from random import randrange

filename = '22.xlsx'
wb = load_workbook(filename)
ws = wb["Лист1"]
CONST = 17 + 1
times = []
while len(times) < 50:
    x = random.randrange(100, 1000, 5)
    if x not in times:
        times.append(x)
ID = [(1, '0')]
while len(ID) < CONST - 1:
    for x in range(2, CONST):
        w = [i for i in list(range(1, CONST)) if i < x]
        random.shuffle(w)
        while len(w) > 2:
            random.shuffle(w)
            w.pop(0)
        w = sorted(w)
        ID.append((x, ''.join(str(x) + ';' if x != w[-1] else str(x) for x in w)))
        if len(ID) == CONST - 1:
            break
print(ID)
print(len(ID))
for x, y in ID:
    ws["C" + str(x + 1)].value = y
    print(x, y)
for x in range(2, CONST + 1):
    ws["A" + str(x)].value = x - 1
    ws["B" + str(x)].value = times[x]

wb.save(filename)
