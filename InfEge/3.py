product = []
with open('file.txt', encoding="utf-8") as t:
    for i in range(60):
        product.append([i.strip() for i in t.readline().split('-')])
# print(product)
from openpyxl import *
import datetime as DT
import random
from random import randrange
from faker import Faker
fake = Faker('ru_RU')
Faker.seed()
streets = [fake.street_address() for i in range(30)]
#print(streets)
import pandas as pd
from datetime import datetime
from datetime import timedelta

DATE = []
CONST = 5001


def generate_range_dates(start_date, end_date) -> list:
    date_1 = min(start_date, end_date)
    date_2 = max(start_date, end_date)

    # Сразу добавляем стартовую дату
    items = [date_1]

    while date_1 < date_2:
        date_1 += DT.timedelta(days=1)
        items.append(date_1)

    return items


def d2s(date):
    return date.strftime("%d.%m.%Y")


start_date = DT.date(2023, 1, 10)
end_date = DT.date(2023, 6, 10)

items = generate_range_dates(start_date, end_date)
stack_data = []
for date in items:
    DATE += [d2s(date)]
for x in range(2, CONST + 1):
    stack_data += [random.choice(DATE)]
stack_data.sort(key=lambda x: (x[3] + x[4], x[0] + x[1]))
# print(stack_data)
art = []
for x in range(1, 20 + 1):
    art += ['M' + str(x)] * 250
art += ['M20'] * 2
wb = load_workbook('test.xlsx')
ws = wb["Движение товаров"]
for x in range(2, CONST + 1):
    ws["A" + str(x)].value = x-1
for x in range(2, len(stack_data)):
    ws["B" + str(x)].value = stack_data[x]
ws["B" + str(CONST)].value = '10.06.2023'
for x in range(2, len(art)):
    ws["C" + str(x)].value = art[x]
for i1, i2 in zip(range(2, CONST + 1), [i for i in range(1, 60 + 1)] * CONST):
    ws["D" + str(i1)].value = i2
for i1, i2 in zip(range(2, CONST + 1), [i for i in range(100, 1000 + 1, 100)] * CONST):
    ws["E" + str(i1)].value = i2
for i1, i2 in zip(range(2, CONST + 1), [i for i in {"Поступление", "Продажа"}] * CONST):
    ws["F" + str(i1)].value = i2
ws = wb["Товар"]
for j, x in enumerate(product, start=1):
    i1, i2, i3, i4, i5 = x
    i4 = randrange(100, 1000, 25)
    i5 = randrange(200, 1000, 100)
    ws["A" + str(j+1)] = j
    ws["B" + str(j + 1)] = i1
    ws["C" + str(j + 1)] = i2
    ws["D" + str(j + 1)] = i3
    ws["E" + str(j + 1)] = i4
    ws["F" + str(j + 1)] = i5
ws = wb["Магазин"]
areas = ['Центральный', 'Промышленный', 'Заречный', 'Промышленный'] * 6
M = ["M" + str(x) for x in range(-1, 20+1)]
for i in range(2, len(M)):
    ws["A" + str(i)] = M[i]
    ws["B" + str(i)] = areas[i]
    ws["C" + str(i)] = streets[i]

wb.save('test.xlsx')
